import os
import sys
import uuid

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from sqlalchemy.orm import Session
from datetime import datetime
from io import BytesIO
from fastapi.responses import StreamingResponse

from database import get_db, Bill
from parser import extract_pdf_text, parse_bill, make_bill_data

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI(title="BillReader API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(HTTPException)
def custom_http_exception_handler(request, exc):
    detail = exc.detail
    if isinstance(detail, dict):
        return JSONResponse(status_code=exc.status_code, content=detail)
    return JSONResponse(
        status_code=exc.status_code,
        content={"success": False, "message": str(detail)},
    )


def bill_to_dict(b: Bill) -> dict:
    result = {}
    for column in b.__table__.columns:
        val = getattr(b, column.name)
        if isinstance(val, datetime):
            result[column.name] = val.isoformat()
        else:
            result[column.name] = val
    return result


def find_duplicate(db: Session, bill_data: dict) -> Bill | None:
    bn = bill_data.get("billNumber")
    mn = bill_data.get("meterNumber")
    bm = bill_data.get("billMonth")
    pa = bill_data.get("payableAmount")
    cb = bill_data.get("currentBill")

    if not bn:
        return None

    existing = db.query(Bill).filter(Bill.billNumber == bn).first()
    if not existing:
        return None

    same_meter = mn and existing.meterNumber == mn
    same_month = bm and existing.billMonth == bm
    same_account = existing.accountNumber == bill_data.get("accountNumber")
    same_payable = pa is not None and existing.payableAmount == pa
    same_bill = cb is not None and existing.currentBill == cb

    if same_meter or same_month or same_account or same_payable or same_bill:
        return existing

    return None


@app.get("/api/bills")
def list_bills(db: Session = Depends(get_db)):
    try:
        bills = (
            db.query(Bill)
            .order_by(Bill.uploadedAt.desc())
            .all()
        )
        data = [bill_to_dict(b) for b in bills]

        total = len(data)
        total_consumption = sum(b.get("consumption") or 0 for b in data)
        total_current_bill = sum(b.get("currentBill") or 0 for b in data)
        total_payable = sum(b.get("payableAmount") or 0 for b in data)

        summary = {
            "totalBills": total,
            "totalConsumption": total_consumption,
            "totalCurrentBill": total_current_bill,
            "totalPayable": total_payable,
            "averageConsumption": total_consumption / total if total else 0,
            "averageBill": total_current_bill / total if total else 0,
        }

        return {"success": True, "summary": summary, "data": data}
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail="Failed to fetch bills.")


@app.post("/api/upload")
async def upload_bill(request: Request, db: Session = Depends(get_db)):
    try:
        content_type = request.headers.get("content-type", "")
        if "multipart/form-data" not in content_type:
            raise HTTPException(status_code=400, detail={"success": False, "message": "No PDF uploaded"})

        try:
            form = await request.form()
        except Exception:
            raise HTTPException(status_code=400, detail={"success": False, "message": "No PDF uploaded"})

        file = form.get("file")

        if not file or not hasattr(file, "filename"):
            raise HTTPException(status_code=400, detail={"success": False, "message": "No PDF uploaded"})

        if not file.filename or not file.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail={"success": False, "message": "Only PDF files are supported"})

        contents = await file.read()
        raw_text = extract_pdf_text(contents)
        bill = parse_bill(raw_text)
        bill_data = make_bill_data(bill)

        dup = find_duplicate(db, bill_data)
        if dup:
            return {
                "success": True,
                "duplicate": True,
                "message": "Bill already exists.",
                "bill": bill_to_dict(dup),
            }

        record = Bill(**bill_data)
        db.add(record)
        db.commit()
        db.refresh(record)

        pdf_filename = f"{record.id}.pdf"
        pdf_path = os.path.join(UPLOAD_DIR, pdf_filename)
        with open(pdf_path, "wb") as f:
            f.write(contents)
        record.pdfPath = pdf_filename
        db.commit()
        db.refresh(record)

        return {
            "success": True,
            "message": "Bill parsed successfully.",
            "data": bill_to_dict(record),
        }
    except HTTPException:
        raise
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail={"success": False, "message": "Failed to parse bill."})


@app.post("/api/upload/batch")
async def upload_bills_batch(request: Request, db: Session = Depends(get_db)):
    try:
        content_type = request.headers.get("content-type", "")
        if "multipart/form-data" not in content_type:
            raise HTTPException(status_code=400, detail={"success": False, "message": "No files uploaded"})

        try:
            form = await request.form()
        except Exception:
            raise HTTPException(status_code=400, detail={"success": False, "message": "No files uploaded"})

        files_list = form.getlist("files")
        if not files_list:
            raise HTTPException(status_code=400, detail={"success": False, "message": "No files uploaded"})

        results = []
        for upload_file in files_list:
            if not upload_file.filename or not upload_file.filename.lower().endswith(".pdf"):
                results.append({
                    "fileName": upload_file.filename or "unknown",
                    "success": False,
                    "message": "Not a PDF file",
                })
                continue

            try:
                contents = await upload_file.read()
                raw_text = extract_pdf_text(contents)
                bill = parse_bill(raw_text)
                bill_data = make_bill_data(bill)

                dup = find_duplicate(db, bill_data)
                if dup:
                    results.append({
                        "fileName": upload_file.filename,
                        "success": True,
                        "duplicate": True,
                        "message": "Bill already exists.",
                        "data": bill_to_dict(dup),
                    })
                    continue

                record = Bill(**bill_data)
                db.add(record)
                db.commit()
                db.refresh(record)

                pdf_filename = f"{record.id}.pdf"
                pdf_path = os.path.join(UPLOAD_DIR, pdf_filename)
                with open(pdf_path, "wb") as f:
                    f.write(contents)
                record.pdfPath = pdf_filename
                db.commit()
                db.refresh(record)

                results.append({
                    "fileName": upload_file.filename,
                    "success": True,
                    "message": "Bill parsed successfully.",
                    "data": bill_to_dict(record),
                })
            except Exception as e:
                print(f"Error processing {upload_file.filename}: {e}")
                results.append({
                    "fileName": upload_file.filename,
                    "success": False,
                    "message": f"Failed to parse: {str(e)}",
                })

        return {"success": True, "results": results}
    except HTTPException:
        raise
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail={"success": False, "message": "Batch upload failed."})


@app.get("/api/bills/export")
def export_bills(accounts: str = "", db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        query = db.query(Bill).order_by(Bill.billMonth.asc())

        account_list = [a.strip() for a in accounts.split(",") if a.strip()]
        if account_list:
            query = query.filter(Bill.accountNumber.in_(account_list))

        all_bills = query.all()

        grouped: dict[str, list[Bill]] = {}
        for b in all_bills:
            key = b.accountNumber or "Unknown"
            grouped.setdefault(key, []).append(b)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        main_columns = [
            ("Bill Month", 14),
            ("Account Number", 18),
            ("Bill Number", 18),
            ("Consumer Name", 22),
            ("Meter Number", 16),
            ("Address", 30),
            ("Division", 16),
            ("Subdivision", 16),
            ("Category", 14),
            ("Connection Type", 16),
            ("Supply Type", 14),
            ("Sanctioned Load (kW)", 16),
            ("Billed Demand (kW)", 16),
            ("Security Deposit (₹)", 18),
            ("Bill Date", 14),
            ("Due Date", 14),
            ("Disconnection Date", 18),
            ("Billing Start", 16),
            ("Billing End", 14),
            ("Connection Date", 16),
            ("Previous Reading (kWh)", 20),
            ("Current Reading (kWh)", 20),
            ("Consumption (kWh)", 18),
            ("Previous Reading (kVAh)", 20),
            ("Current Reading (kVAh)", 20),
            ("Consumption (kVAh)", 18),
            ("Power Factor", 14),
            ("Solar Export (kWh)", 18),
            ("Opening Solar Balance", 20),
            ("Closing Solar Balance", 20),
            ("Current Bill (₹)", 16),
            ("Payable Amount (₹)", 18),
            ("Previous Due (₹)", 16),
            ("Energy Charges (₹)", 18),
            ("Demand Charges (₹)", 18),
            ("Electricity Duty (₹)", 20),
            ("FPPA Surcharge (₹)", 18),
            ("Minimum Charges (₹)", 18),
            ("Excess Demand Penalty (₹)", 24),
            ("Other Charges (₹)", 18),
            ("Subsidy (₹)", 14),
            ("Arrears (₹)", 14),
        ]

        extra_columns = [
            ("Credit (₹)", 14),
            ("Debit (₹)", 14),
            ("Rebate (₹)", 14),
            ("Meter Charges (₹)", 18),
            ("Due Security (₹)", 16),
            ("Payment Date", 14),
            ("Payment Amount (₹)", 18),
            ("Payment Mode", 16),
            ("Receipt Number", 16),
            ("Father Name", 18),
            ("Raw Text", 50),
        ]

        def fmt_dt(dt):
            return dt.isoformat() if dt else None

        def fmt_num(v):
            return float(v) if v is not None else None

        all_columns = main_columns + extra_columns

        for account, bills in grouped.items():
            ws = wb.create_sheet(title=str(account)[:31])

            for col_idx, (col_name, _) in enumerate(all_columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row_idx, bill in enumerate(bills, 2):
                vals = [
                    bill.billMonth,
                    bill.accountNumber,
                    bill.billNumber,
                    bill.consumerName,
                    bill.meterNumber,
                    bill.address,
                    bill.division,
                    bill.subdivision,
                    bill.category,
                    bill.connectionType,
                    bill.supplyType,
                    bill.sanctionedLoad,
                    bill.billedDemand,
                    bill.securityDeposit,
                    fmt_dt(bill.billDate),
                    fmt_dt(bill.dueDate),
                    fmt_dt(bill.disconnectionDate),
                    fmt_dt(bill.billingStart),
                    fmt_dt(bill.billingEnd),
                    fmt_dt(bill.connectionDate),
                    fmt_num(bill.previousReading),
                    fmt_num(bill.currentReading),
                    fmt_num(bill.consumption),
                    fmt_num(bill.previousKVAH),
                    fmt_num(bill.currentKVAH),
                    fmt_num(bill.kvaConsumption),
                    bill.powerFactor,
                    fmt_num(bill.solarExport),
                    fmt_num(bill.openingSolarBalance),
                    fmt_num(bill.closingSolarBalance),
                    fmt_num(bill.currentBill),
                    fmt_num(bill.payableAmount),
                    fmt_num(bill.previousDue),
                    fmt_num(bill.energyCharges),
                    fmt_num(bill.demandCharges),
                    fmt_num(bill.electricityDuty),
                    fmt_num(bill.fppa),
                    fmt_num(bill.minimumCharges),
                    fmt_num(bill.excessDemandPenalty),
                    fmt_num(bill.otherCharges),
                    fmt_num(bill.subsidy),
                    fmt_num(bill.arrears),
                ]
                # extra columns
                extra_vals = [
                    fmt_num(bill.credit),
                    fmt_num(bill.debit),
                    fmt_num(bill.rebate),
                    fmt_num(bill.meterCharges),
                    fmt_num(bill.dueSecurity),
                    fmt_dt(bill.paymentDate),
                    fmt_num(bill.paymentAmount),
                    bill.paymentMode,
                    bill.receiptNumber,
                    bill.fatherName,
                    bill.rawText,
                ]
                for col_idx, val in enumerate(vals + extra_vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border

            for col_idx, (_, width) in enumerate(all_columns, 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=bills.xlsx"},
        )
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail={"success": False, "message": "Failed to export bills."})


@app.get("/api/bills/{bill_id}/pdf")
def get_bill_pdf(bill_id: str, db: Session = Depends(get_db)):
    try:
        bill = db.query(Bill).filter(Bill.id == bill_id).first()
        if not bill:
            raise HTTPException(status_code=404, detail={"success": False, "message": "Bill not found"})
        if not bill.pdfPath:
            raise HTTPException(status_code=404, detail={"success": False, "message": "PDF not available for this bill"})
        pdf_path = os.path.join(UPLOAD_DIR, bill.pdfPath)
        if not os.path.exists(pdf_path):
            raise HTTPException(status_code=404, detail={"success": False, "message": "PDF file not found on disk"})
        return FileResponse(pdf_path, media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="{bill.billNumber}.pdf"'})
    except HTTPException:
        raise
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail={"success": False, "message": "Failed to retrieve PDF."})


@app.delete("/api/bills")
def delete_all_bills(db: Session = Depends(get_db)):
    try:
        bills = db.query(Bill).all()
        for bill in bills:
            if bill.pdfPath:
                pdf_path = os.path.join(UPLOAD_DIR, bill.pdfPath)
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
        count = db.query(Bill).delete()
        db.commit()
        return {"success": True, "message": f"{count} bill(s) deleted."}
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail={"success": False, "message": "Failed to delete bills."})


@app.delete("/api/bills/{bill_id}")
def delete_bill(bill_id: str, db: Session = Depends(get_db)):
    try:
        bill = db.query(Bill).filter(Bill.id == bill_id).first()
        if not bill:
            raise HTTPException(status_code=404, detail={"success": False, "message": "Bill not found"})
        if bill.pdfPath:
            pdf_path = os.path.join(UPLOAD_DIR, bill.pdfPath)
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
        db.delete(bill)
        db.commit()
        return {"success": True, "message": "Bill deleted."}
    except HTTPException:
        raise
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail={"success": False, "message": "Failed to delete bill."})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
